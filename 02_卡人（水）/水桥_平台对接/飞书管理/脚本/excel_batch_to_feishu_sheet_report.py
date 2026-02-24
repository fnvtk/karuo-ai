#!/usr/bin/env python3
"""
批量 Excel/CSV → 飞书电子表格（Sheets）写入，并生成日报图表（图片）+ 发飞书群。

设计目标（最小可用）：
1) 支持目录/多文件批量导入（xlsx/xlsm/csv）
2) 写入到同一个 spreadsheet 的同一张 sheet（按块堆叠，每个文件一个区块）
3) 生成本地日报：summary + 折线图（最多3列）→ HTML → PNG
4) 通过飞书群机器人 webhook 发送：文字 + 图片

注意：
- 写表使用「用户 access_token」（来自同目录 .feishu_tokens.json）
- 若 token 过期：优先提示先跑 auto_log.py（已有静默刷新机制）
- 如需脚本内自动 refresh：请配置环境变量 FEISHU_APP_ID / FEISHU_APP_SECRET
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, List, Optional, Tuple
from urllib.parse import quote

import requests

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:
    load_workbook = None  # venv 未安装时给出明确提示

try:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt  # type: ignore
except Exception:
    plt = None


SCRIPT_DIR = Path(__file__).resolve().parent
TOKEN_FILE = SCRIPT_DIR / ".feishu_tokens.json"

# 输出目录（按卡若AI 输出规范）
OUT_ROOT = Path("/Users/karuo/Documents/卡若Ai的文件夹/导出/飞书/Excel日报")


@dataclass
class FeishuAuth:
    access_token: str
    refresh_token: str | None = None


def _now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H%M%S")


def _to_text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime,)):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    return str(v)


def _safe_title(name: str) -> str:
    base = re.sub(r"\s+", " ", name.strip())
    base = re.sub(r"[^\w\u4e00-\u9fff\-\.\(\) ]+", "_", base)
    return base[:80] or "未命名"


def load_user_tokens() -> FeishuAuth:
    if not TOKEN_FILE.exists():
        raise RuntimeError(
            f"未找到飞书 token 文件：{TOKEN_FILE}\n"
            f"请先运行：python3 {SCRIPT_DIR}/auto_log.py 完成授权/刷新"
        )
    data = json.loads(TOKEN_FILE.read_text(encoding="utf-8"))
    access = (data.get("access_token") or "").strip()
    refresh = (data.get("refresh_token") or "").strip() or None
    if not access:
        raise RuntimeError("token 文件存在，但缺少 access_token；请先跑 auto_log.py 刷新。")
    return FeishuAuth(access_token=access, refresh_token=refresh)


def try_refresh_access_token(auth: FeishuAuth) -> Optional[str]:
    """
    尝试 refresh（需要 FEISHU_APP_ID/FEISHU_APP_SECRET）。
    不内置默认 app_id/app_secret，避免把密钥扩散到新脚本。
    """
    if not auth.refresh_token:
        return None
    app_id = (os.environ.get("FEISHU_APP_ID") or "").strip()
    app_secret = (os.environ.get("FEISHU_APP_SECRET") or "").strip()
    if not app_id or not app_secret:
        return None

    r = requests.post(
        "https://open.feishu.cn/open-apis/auth/v3/app_access_token/internal",
        json={"app_id": app_id, "app_secret": app_secret},
        timeout=15,
    )
    app_token = (r.json() or {}).get("app_access_token")
    if not app_token:
        return None

    r2 = requests.post(
        "https://open.feishu.cn/open-apis/authen/v1/oidc/refresh_access_token",
        headers={"Authorization": f"Bearer {app_token}", "Content-Type": "application/json"},
        json={"grant_type": "refresh_token", "refresh_token": auth.refresh_token},
        timeout=15,
    )
    out = r2.json() or {}
    if out.get("code") != 0:
        return None

    data = out.get("data") or {}
    new_access = (data.get("access_token") or "").strip()
    new_refresh = (data.get("refresh_token") or "").strip() or auth.refresh_token
    if not new_access:
        return None

    # 写回 token 文件（沿用旧结构）
    old = json.loads(TOKEN_FILE.read_text(encoding="utf-8"))
    old["access_token"] = new_access
    old["refresh_token"] = new_refresh
    TOKEN_FILE.write_text(json.dumps(old, ensure_ascii=False, indent=2), encoding="utf-8")
    return new_access


def feishu_put_values(
    access_token: str,
    spreadsheet_token: str,
    range_str: str,
    values: List[List[Any]],
    value_input_option: str = "RAW",
) -> Tuple[int, dict]:
    url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values"
    params = {"valueInputOption": value_input_option}
    payload = {"valueRange": {"range": range_str, "values": values}}
    r = requests.put(
        url,
        params=params,
        headers={"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"},
        json=payload,
        timeout=30,
    )
    try:
        return r.status_code, r.json()
    except Exception:
        return r.status_code, {"code": -1, "msg": (r.text or "")[:200]}


def feishu_read_values(
    access_token: str,
    spreadsheet_token: str,
    range_str: str,
) -> Tuple[int, dict]:
    url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}"
        f"/values/{quote(range_str, safe='')}"
    )
    r = requests.get(
        url,
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=30,
    )
    try:
        return r.status_code, r.json()
    except Exception:
        return r.status_code, {"code": -1, "msg": (r.text or "")[:200]}


def _col_letter(n: int) -> str:
    # 0->A, 1->B ... 25->Z, 26->AA
    s = ""
    while True:
        s = chr(65 + n % 26) + s
        n = n // 26
        if n <= 0:
            break
    return s


def read_csv(path: Path, max_rows: int, max_cols: int) -> List[List[Any]]:
    rows: List[List[Any]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append([c for c in row[:max_cols]])
    return rows


def read_xlsx(path: Path, worksheet: Optional[str], max_rows: int, max_cols: int) -> List[List[Any]]:
    if load_workbook is None:
        raise RuntimeError(
            "未安装 openpyxl（或依赖异常）。请使用 venv 运行：\n"
            f'  "/Users/karuo/.venvs/karuo-feishu/bin/python" "{path}" ...'
        )
    wb = load_workbook(path, data_only=True, read_only=True)
    if worksheet:
        ws = wb[worksheet]
    else:
        ws = wb.active
    rows: List[List[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        out_row: List[Any] = []
        for j, v in enumerate(row):
            if j >= max_cols:
                break
            out_row.append("" if v is None else v)
        rows.append(out_row)
    return rows


def normalize_rows(rows: List[List[Any]]) -> List[List[Any]]:
    max_len = max((len(r) for r in rows), default=0)
    norm: List[List[Any]] = []
    for r in rows:
        rr = list(r) + [""] * (max_len - len(r))
        norm.append([_to_text(v) if isinstance(v, (datetime,)) else ("" if v is None else v) for v in rr])
    return norm


def detect_numeric_columns(table: List[List[Any]], has_header: bool) -> List[Tuple[int, str, List[float]]]:
    if not table:
        return []
    start = 1 if has_header else 0
    headers = [f"列{idx+1}" for idx in range(len(table[0]))]
    if has_header:
        headers = [str(c).strip() if str(c).strip() else f"列{idx+1}" for idx, c in enumerate(table[0])]

    numeric_cols: List[Tuple[int, str, List[float]]] = []
    for ci in range(len(headers)):
        vals: List[float] = []
        for r in table[start:]:
            if ci >= len(r):
                continue
            v = r[ci]
            if v is None or v == "":
                continue
            try:
                vals.append(float(v))
            except Exception:
                # 非数字列跳过
                vals = []
                break
        if len(vals) >= 3:
            numeric_cols.append((ci, headers[ci], vals))
    return numeric_cols


def plot_numeric_columns(
    table: List[List[Any]],
    has_header: bool,
    out_dir: Path,
    top_k: int = 3,
) -> List[Path]:
    if plt is None:
        return []
    cols = detect_numeric_columns(table, has_header)
    if not cols:
        return []

    # 简单选择：按有效值数量降序取前 top_k
    cols.sort(key=lambda x: len(x[2]), reverse=True)
    selected = cols[:top_k]

    img_paths: List[Path] = []
    for idx, (ci, name, vals) in enumerate(selected, start=1):
        fig = plt.figure(figsize=(10, 3))
        ax = fig.add_subplot(111)
        ax.plot(list(range(1, len(vals) + 1)), vals, linewidth=2)
        ax.set_title(name)
        ax.set_xlabel("行序号")
        ax.set_ylabel("数值")
        ax.grid(True, alpha=0.3)
        img = out_dir / f"chart_{idx}_{_safe_title(name)}.png"
        fig.tight_layout()
        fig.savefig(img, dpi=160)
        plt.close(fig)
        img_paths.append(img)
    return img_paths


def build_report_html(
    title: str,
    sections: List[Tuple[str, List[str], List[Path]]],
    out_html: Path,
) -> None:
    parts: List[str] = []
    parts.append("<!doctype html><html><head><meta charset='utf-8'/>")
    parts.append("<style>")
    parts.append(
        "body{font-family:-apple-system,BlinkMacSystemFont,'PingFang SC',sans-serif;"
        "background:#0b1220;color:#e8eefc;margin:0;padding:24px}"
    )
    parts.append(".card{background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.12);"
                 "border-radius:16px;padding:18px;margin:14px 0}")
    parts.append("h1{margin:0 0 6px 0;font-size:22px}")
    parts.append("h2{margin:0 0 10px 0;font-size:18px}")
    parts.append("p,li{line-height:1.5;color:#cfe0ff}")
    parts.append("img{max-width:100%;border-radius:12px;border:1px solid rgba(255,255,255,0.14)}")
    parts.append(".muted{color:#9bb3db;font-size:12px}")
    parts.append("</style></head><body>")

    parts.append(f"<div class='card'><h1>{title}</h1>")
    parts.append(f"<div class='muted'>生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div></div>")

    for sec_title, bullets, imgs in sections:
        parts.append("<div class='card'>")
        parts.append(f"<h2>{sec_title}</h2>")
        if bullets:
            parts.append("<ul>")
            for b in bullets:
                parts.append(f"<li>{b}</li>")
            parts.append("</ul>")
        for img in imgs:
            parts.append(f"<p><img src='{img.name}'/></p>")
        parts.append("</div>")

    parts.append("</body></html>")
    out_html.write_text("\n".join(parts), encoding="utf-8")


def screenshot_html_to_png(html_path: Path, png_path: Path, width: int = 1200) -> None:
    # 直接用 playwright（已在 venv 安装）
    from playwright.sync_api import sync_playwright  # type: ignore

    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page(viewport={"width": width, "height": 800})
        page.goto(f"file://{html_path.absolute()}")
        page.wait_for_timeout(700)
        height = page.evaluate("document.body.scrollHeight")
        page.set_viewport_size({"width": width, "height": int(height)})
        page.screenshot(path=str(png_path), full_page=True)
        browser.close()


def send_feishu_text(webhook: str, text: str) -> None:
    payload = {"msg_type": "text", "content": {"text": text}}
    r = requests.post(webhook, json=payload, timeout=15)
    try:
        out = r.json()
    except Exception:
        out = {"code": -1, "msg": (r.text or "")[:200]}
    if out.get("code") != 0:
        raise RuntimeError(f"飞书群消息发送失败：{out.get('msg') or out}")


def send_feishu_image_via_tool(python_bin: str, image_path: Path, webhook: str) -> None:
    """
    复用「智能纪要」的图片上传+发群工具，避免在本脚本内复制 upload 逻辑。
    """
    tool = Path("/Users/karuo/Documents/个人/卡若AI/02_卡人（水）/水桥_平台对接/智能纪要/脚本/send_to_feishu.py")
    if not tool.exists():
        raise RuntimeError(f"未找到发送工具：{tool}")
    # 由 tool 内部处理 tenant_access_token + 上传图片 + 发 image 消息
    import subprocess

    subprocess.check_call([python_bin, str(tool), "--image", str(image_path), "--webhook", webhook])


def main() -> None:
    parser = argparse.ArgumentParser(description="批量 Excel/CSV → 飞书表格 + 日报图表 + 发群")
    parser.add_argument("--input", "-i", required=True, help="文件或目录（支持 .xlsx/.xlsm/.csv）")
    parser.add_argument("--spreadsheet-token", default=os.environ.get("FEISHU_SPREADSHEET_TOKEN", "").strip(),
                        help="飞书 spreadsheet_token（必填，可用环境变量 FEISHU_SPREADSHEET_TOKEN）")
    parser.add_argument("--sheet-id", default=os.environ.get("FEISHU_SHEET_ID", "").strip(),
                        help="飞书 sheet_id（必填，可用环境变量 FEISHU_SHEET_ID）")
    parser.add_argument("--start-row", type=int, default=1, help="写入起始行（默认1，即A1）")
    parser.add_argument("--gap-rows", type=int, default=2, help="每个文件区块之间空几行（默认2）")
    parser.add_argument("--max-rows", type=int, default=800, help="每个文件最多读取多少行（默认800）")
    parser.add_argument("--max-cols", type=int, default=40, help="每个文件最多读取多少列（默认40）")
    parser.add_argument("--worksheet", default="", help="xlsx 指定工作表名（默认 active）")
    parser.add_argument("--no-header", action="store_true", help="首行不是表头（默认首行为表头）")
    parser.add_argument("--webhook", default=os.environ.get("FEISHU_GROUP_WEBHOOK", "").strip(),
                        help="飞书群机器人 webhook（可用环境变量 FEISHU_GROUP_WEBHOOK）")
    parser.add_argument("--send-image", action="store_true", help="截图日报并发送图片（默认只发文字）")
    parser.add_argument("--title", default="", help="日报标题（可选）")
    parser.add_argument("--python-bin", default="/Users/karuo/.venvs/karuo-feishu/bin/python",
                        help="用于发送图片的 python（默认使用 karuo-feishu venv）")
    args = parser.parse_args()

    if not args.spreadsheet_token or not args.sheet_id:
        raise SystemExit("缺少 spreadsheet-token 或 sheet-id（可通过环境变量传入）。")
    if not args.webhook:
        raise SystemExit("缺少 webhook（可通过 --webhook 或 FEISHU_GROUP_WEBHOOK）。")

    input_path = Path(args.input).expanduser().resolve()
    if input_path.is_dir():
        files = sorted([p for p in input_path.iterdir() if p.suffix.lower() in (".xlsx", ".xlsm", ".csv")])
    else:
        files = [input_path]
    if not files:
        raise SystemExit("未找到可导入文件（xlsx/xlsm/csv）。")

    auth = load_user_tokens()
    access_token = auth.access_token

    # 输出目录
    run_dir = OUT_ROOT / f"{datetime.now().strftime('%Y-%m-%d')}_{_now_str()}"
    charts_dir = run_dir / "charts"
    run_dir.mkdir(parents=True, exist_ok=True)
    charts_dir.mkdir(parents=True, exist_ok=True)

    has_header = not args.no_header
    worksheet = args.worksheet.strip() or None

    # 逐文件写入（按块堆叠在同一 sheet）
    sections: List[Tuple[str, List[str], List[Path]]] = []
    cur_row = max(1, int(args.start_row))

    for fp in files:
        ext = fp.suffix.lower()
        if ext == ".csv":
            raw = read_csv(fp, args.max_rows, args.max_cols)
        else:
            raw = read_xlsx(fp, worksheet, args.max_rows, args.max_cols)
        table = normalize_rows(raw)

        title = _safe_title(fp.stem)
        if not table:
            sections.append((title, ["空表：无可写入数据"], []))
            continue

        rows_n = len(table)
        cols_n = max((len(r) for r in table), default=0)

        # 先写一个区块标题行
        header_row = [[f"=== {title} ==="]]
        rng_title = f"{args.sheet_id}!A{cur_row}:A{cur_row}"
        code, body = feishu_put_values(access_token, args.spreadsheet_token, rng_title, header_row, "RAW")
        if code == 401 or body.get("code") in (99991677, 99991663):
            new_access = try_refresh_access_token(auth)
            if new_access:
                access_token = new_access
                code, body = feishu_put_values(access_token, args.spreadsheet_token, rng_title, header_row, "RAW")
        if code != 200 or body.get("code") not in (0, None):
            raise RuntimeError(f"写入标题行失败：{code} {body}")

        cur_row += 1

        # 写入表格
        end_col_letter = _col_letter(max(cols_n - 1, 0))
        rng = f"{args.sheet_id}!A{cur_row}:{end_col_letter}{cur_row + rows_n - 1}"
        code, body = feishu_put_values(access_token, args.spreadsheet_token, rng, table, "RAW")
        if code == 401 or body.get("code") in (99991677, 99991663):
            new_access = try_refresh_access_token(auth)
            if new_access:
                access_token = new_access
                code, body = feishu_put_values(access_token, args.spreadsheet_token, rng, table, "RAW")
        if code != 200 or body.get("code") not in (0, None):
            raise RuntimeError(f"写入表格失败：{fp.name} / {code} {body}")

        # 生成图表
        file_out_dir = charts_dir / title
        file_out_dir.mkdir(parents=True, exist_ok=True)
        charts = plot_numeric_columns(table, has_header, file_out_dir, top_k=3)

        bullets = [
            f"文件：{fp.name}",
            f"写入范围：{rng}",
            f"行数：{rows_n}；列数：{cols_n}",
        ]
        if charts:
            bullets.append(f"已生成图表：{len(charts)} 张（最多3列）")
        else:
            bullets.append("未生成图表：缺少可识别的数字列（或未安装 matplotlib）")

        sections.append((title, bullets, charts))

        cur_row = cur_row + rows_n + int(args.gap_rows)

    report_title = args.title.strip() or f"飞书表格日报（Excel批量导入）"
    html = run_dir / "report.html"
    png = run_dir / "report.png"

    # HTML 里引用图片，确保与 html 同目录可相对加载
    # 将 charts 复制/软链到同目录（这里用复制路径写入时只引用文件名，因此需要都在同目录）
    # 简化：把图片都放在 run_dir 下的 charts_flat/，并在 HTML 中用相对路径。
    charts_flat = run_dir / "charts_flat"
    charts_flat.mkdir(parents=True, exist_ok=True)

    new_sections: List[Tuple[str, List[str], List[Path]]] = []
    for sec_title, bullets, imgs in sections:
        flat_imgs: List[Path] = []
        for img in imgs:
            target = charts_flat / f"{sec_title}__{img.name}"
            if not target.exists():
                target.write_bytes(img.read_bytes())
            flat_imgs.append(target)
        new_sections.append((sec_title, bullets, flat_imgs))

    build_report_html(report_title, new_sections, html)
    screenshot_html_to_png(html, png, width=1200)

    # 发群：文字 +（可选）图片
    msg_lines = [
        "【Excel→飞书表格】批量写入完成",
        f"spreadsheet_token: {args.spreadsheet_token}",
        f"sheet_id: {args.sheet_id}",
        f"文件数：{len(files)}",
        f"日报：{png}",
    ]
    send_feishu_text(args.webhook, "\n".join(msg_lines))

    if args.send_image:
        send_feishu_image_via_tool(args.python_bin, png, args.webhook)

    print(f"✅ 写入完成：{len(files)} 个文件")
    print(f"✅ 日报输出：{png}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        raise
    except Exception as e:
        print(f"❌ 失败：{e}")
        sys.exit(1)

